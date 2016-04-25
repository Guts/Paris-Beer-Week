# -*- coding: UTF-8 -*-
#!/usr/bin/env python3.4
from __future__ import (unicode_literals, print_function)
# -----------------------------------------------------------------------------
# Name:         Parser Geocoder from Excel file
# Purpose:      todo
#
# Author:       Julien Moura (https://twitter.com/geojulien)
#
# Python:       2.7.x
# Created:      14/05/2015
# Updated:      15/05/2015
#
# Licence:      GPL 3
# ------------------------------------------------------------------------------

# #############################################################################
# ########## Libraries #############
# ##################################

# Standard library
from calendar import timegm
from datetime import datetime
import locale
from time import mktime

# Python 3 backported

# 3rd party libraries
from openpyxl import load_workbook
from geojson import dump, Feature, FeatureCollection, Point
import pytz

# Custom modules

# #############################################################################
# ######### Main program ##########
# #################################

edition = 2016

# ############################# PARTICIPANTS

# Structure attendue ##################################
# col_idx   col_name        description
# 0         ID
# 1         NOM
# 2         TYPE
# 3         DESCR_FR
# 4         DESCR_EN
# 5         WEBSITE
# 6         ADR_NUM
# 7         ADR_CMPLT
# 8         ADR_TYP
# 9         ADR_LIB
# 10        ADR_CP
# 11        ADR_CITY
# 12        ADR_COUNTRY
# 13        ADR_CONCAT
# 14        TEL
# 15        OSM_URL
# 16        GMAPS_URL
# 17        ed_01_2014
# 18        ed_02_2015
# 19        ed_03_2016
# 20        ed14_URL_FR
# 21        ed15_URL_FR
# 22        ed15_URL_EN
# 23        ed16_URL_FR
# 24        ed16_URL_EN
# 25        URL_INSTA
# 26        URL_FB
# 27        URL_TWITTER
# 28        URL_GPLUS
# 29        URL_THUMB
# 30        URL_MAIL
# 31        x_longitude
# 32        y_latitude
# 33        X_NOMINATIM
# 34        Y_NOMINATIM
# 35        LI_ID_EVT
# 36        URL_CITYMAPPER

# /Structure attendue ##################################

# liste pour stocker les objets
li_objs = []

# ouverture du fichier des participants en lecture
wb = load_workbook(filename='ParisBeerWeek_participants_{}.xlsx'.format(edition),
                   read_only=True,
                   guess_types=True,
                   data_only=True,
                   use_iterators=True
                   )

# noms des onglets
# print(wb.get_sheet_names())

ws = wb.worksheets[0]  # ws = première feuille

row_count = ws.max_row
column_count = ws.max_column

print(row_count)
print(column_count)

for row in ws.iter_rows(row_offset=1):
    # extraire le nom
    nom = row[1].value

    # vérification qu'il s'agit bien d'une ligne remplie
    if not nom:
        print('\nFin du tableau')
        break
    else:
        pass

    # si l'adresse n'est pas renseignée, on s'arrache
    if not row[26].value and not row[27].value:
        print('\nCoordinates NR, use Geocoder before: ' + str(row[0].value))
        continue
    else:
        pass

    # extraire l'adresse
    # libelle = str(row[6].value) + " " + row[8].value + " " + row[9].value
    # ville = row[11].value
    # addr = nom + ", " + ville + ", France"
    addr = row[13].value

    # extraction des coordonnées
    print(row[31].value)
    longitude = row[32].value
    latitude = row[31].value

    # 
    if not longitude or not latitude:
        continue
    else:
        pass

    # extraction du téléphone
    if row[14].value:
        tel = row[14].value
    else:
        tel = "NR"

    # stockage dans des objets en vue de la sérialisation
    point = Point((longitude, latitude))
    obj = Feature(geometry=point,
                  id=row[0].value,
                  properties={"ADR_COUNTRY": row[12].value,
                              "NAME": nom,
                              "TYPE": row[2].value,
                              "DESCR_FR": row[3].value,
                              "DESCR_EN": row[4].value,
                              "ADDRESS": addr,
                              "TEL": tel,
                              "WEBSITE": row[5].value,
                              "FACEBOOK": row[22].value,
                              "TWITTER": row[23].value,
                              "PBW_16_FR": row[23].value,
                              "PBW_16_EN": row[24].value,
                              "THUMBNAIL": row[29].value,
                              "OSM": row[15].value,
                              "GMAPS": row[16].value,
                              "CITYMAPPER": row[36].value,
                              "FS_ID": row[37].value,
                              "UTP_ID": row[38].value
                              })
    li_objs.append(obj)

# sérialisation en GeoJSON
featColl = FeatureCollection(li_objs)
with open("../ParisBeerWeek_participants.geojson", "w") as outfile:
    dump(featColl, outfile, sort_keys=True)


# ############################## EVENEMENTS

# Structure attendue ##################################
# col_idx   col_name        description
# 0         ID
# 1         EVT_NOM
# 2         TYPE
# 3         DESCR_FR
# 4         DESCR_EN
# 5         ADR_ID
# 6         ADR_NOM
# 7         ADR_PARTIC
# 8         ADR_NUM
# 9         ADR_COMPL
# 10        ADR_TYP
# 11        ADR_LIB
# 12        ADR_CP
# 13        ADR_CITY
# 14        ADR_COUNTRY
# 15        ADR_CONCAT
# 16        ADR_IMG
# 17        DDAY_START
# 18        TIME_START
# 19        DDAY_END
# 20        TIME_END
# 21        DTIME_START
# 22        DTIME_END
# 23        DUREE
# 24        OSM_URL
# 25        GMAPS_URL
# 26        DDAY_URL_FR
# 27        DDAY_URL_EN
# 28        LI_ID_PART
# 29        ED_YEAR
# 30        X_LONGITUDE
# 31        Y_LATITUDE
# 32        X_NOMINATIM
# 33        Y_NOMINATIM


# liste pour stocker les objets
li_objs = []

# ouverture du fichier des participants en lecture
wb = load_workbook(filename='ParisBeerWeek_evenements_{}.xlsx'.format(edition),
                   read_only=True,
                   guess_types=True,
                   data_only=True,
                   use_iterators=True)

# noms des onglets
# print(wb.get_sheet_names())

ws = wb.worksheets[0]  # ws = première feuille

row_count = ws.get_highest_row() - 1
column_count = ws.get_highest_column() + 1

print(row_count)
print(column_count)

x = 1

for row in ws.iter_rows(row_offset=1):
    # extraire l'adresse
    nom = row[6].value

    # vérification qu'il s'agit bien d'une ligne remplie
    if not nom:
        print('\nFin du tableau')
        break
    else:
        pass

    # si l'adresse n'est pas renseignée, on s'arrache
    if not row[30].value and not row[31].value:
        print('\nCoordinates NR, use Geocoder before: ' + str(row[0].value))
        continue
    else:
        pass

    # extraire l'adresse
    if row[13].value:
        addr = nom + ", " + row[15].value
    else:
        addr = nom + ", Île-de-France, France"

    # date et horaires de l'événement
    paris_tz = pytz.timezone("Europe/Paris")  # fuseau horaire parisien

    locale.setlocale(locale.LC_ALL, '')

    # date et heure de début
    if type(row[21].value) is str:
        evt_start_input = datetime.strptime(row[21].value, "%d/%m/%Y %H:%M:%S")
    else:
        evt_start_input = row[21].value
        pass
    evt_start_input = paris_tz.localize(evt_start_input)
    evt_start_epc = timegm(evt_start_input.timetuple())
    evt_start_txt = evt_start_input.strftime('%A %d %B %Y à %H:%M'.encode('UTF-8'))
    evt_day_txt = evt_start_input.strftime('%A %d %B %Y'.encode('UTF-8'))
    evt_start_time_txt = evt_start_input.strftime('%H:%M'.encode('UTF-8'))
    startDate = evt_start_input.strftime('%d/%m/%Y %H:%M'.encode('UTF-8'))

    # date et heure de fin
    if type(row[22].value) is str:
        evt_end_input = datetime.strptime(row[22].value, "%d/%m/%Y %H:%M:%S")
    else:
        evt_end_input = row[22].value
        pass
    evt_end_input = paris_tz.localize(evt_end_input)
    evt_end_epc = timegm(evt_end_input.timetuple())
    evt_end_txt = evt_end_input.strftime('%A %d %B %Y à %H:%M'.encode('UTF-8'))
    evt_end_time_txt = evt_end_input.strftime('%H:%M'.encode('UTF-8'))
    endDate = evt_end_input.strftime('%d/%m/%Y %H:%M'.encode('UTF-8'))

    # extraction des coordonnées
    longitude = row[30].value
    latitude = row[31].value

    # stockage dans des objets en vue de la sérialisation
    point = Point((longitude, latitude))
    obj = Feature(geometry=point,
                  id=x,
                  properties={"NAME": row[1].value,
                              "DESCR_FR": row[3].value,
                              "DESCR_EN": row[4].value,
                              "ADDRESS": addr,
                              "PLACE_NAME": nom,
                              "EVT_S_TXT": evt_start_txt,
                              "EVT_S_EPC": evt_start_epc,
                              "startDate": startDate,
                              "EVT_S_TIME": evt_start_time_txt,
                              "EVT_E_TXT": evt_end_txt,
                              "EVT_E_EPC": evt_end_epc,
                              "endDate": endDate,
                              "EVT_E_TIME": evt_end_time_txt,
                              "EVT_DDAY": evt_day_txt,
                              "URL_D15_FR": row[26].value,
                              "URL_D15_EN": row[27].value,
                              "OSM": row[24].value,
                              "GMAPS": row[25].value
                              })
    li_objs.append(obj)

    # increment ID
    x += 1


# sérialisation en GeoJSON
featColl = FeatureCollection(li_objs)
with open("../ParisBeerWeek_evenements.geojson", "w") as outfile:
    dump(featColl, outfile, sort_keys=True)


# #############################################################################
# ##### Stand alone program ########
# ##################################

# if __name__ == '__main__':
#     """ standalone execution """
