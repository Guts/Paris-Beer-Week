# -*- coding: UTF-8 -*-
#!/usr/bin/env python3.4
from __future__ import (unicode_literals, print_function)

# ----------------------------------------------------------------------------
# Name:         Parser Geocoder from Excel file
# Purpose:      todo
#
# Author:       Julien Moura (https://twitter.com/geojulien)
#
# Python:       2.7.x
# Created:      14/05/2015
# Updated:      15/05/2015
#
# Licence:      GPL 3
# -----------------------------------------------------------------------------

# ############################################################################
# ########## Libraries #############
# ##################################

# Standard library

# Python 3 backported

# 3rd party libraries
from openpyxl import load_workbook
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut

# Custom modules

# ############################################################################
# ########## Main program ##########
# ##################################

# PARTICIPANTS

# Structure attendue ##################################
# col_idx   col_name        description
# 0         ID
# 1         NOM
# 2         TYPE
# 3         DESCR_FR
# 4         DESCR_EN
# 5         WEBSITE
# 6         ADR_NUM
# 7         ADR_CMPLT
# 8         ADR_TYP
# 9         ADR_LIB
# 10        ADR_CP
# 11        ADR_CITY
# 12        ADR_COUNTRY
# 13        ADR_CONCAT
# 14        TEL
# 15        OSM_URL
# 16        GMAPS_URL
# 17        ed_01_2014
# 18        ed_02_2015
# 19        ed_03_2016
# 20        ed14_URL_FR
# 21        ed15_URL_FR
# 22        ed15_URL_EN
# 23        ed16_URL_FR
# 24        ed16_URL_EN
# 25        URL_INSTA
# 26        URL_FB
# 27        URL_TWITTER
# 28        URL_GPLUS
# 29        URL_THUMB
# 30        URL_MAIL
# 31        x_longitude
# 32        y_latitude
# 33        X_NOMINATIM
# 34        Y_NOMINATIM
# 35        LI_ID_EVT
# 36        URL_CITYMAPPER
# 37        FS_ID
# 38        UTP_ID

# /Structure attendue ##################################

# ouverture du fichier des participants en lecture
wb = load_workbook(filename='ParisBeerWeek_participants.xlsx',
                   # read_only=True,
                   guess_types=True,
                   data_only=True,
                   # use_iterators=True
                   )

# noms des onglets
# print(wb.get_sheet_names())

ws = wb.worksheets[0]  # ws = première feuille

row_count = ws.max_row
column_count = ws.max_column

print(row_count)
print(column_count)

for row in ws.iter_rows(row_offset=1):
    # extraire le nom
    nom = row[1].value

    # vérification qu'il s'agit bien d'une ligne remplie
    if not nom:
        print('\nFin du tableau')
        break
    else:
        pass

    # si l'adresse n'est pas renseignée, on s'arrache
    if not row[8].value and not row[11].value:
        print('\nAdresse NR: ' + str(row[0].value))
        continue
    else:
        pass

    # extraire l'adresse
    # libelle = str(row[6].value) + " " + row[8].value + " " + row[9].value
    ville = row[11].value
    addr = nom + ", " + ville + ", France"
    addr2 = row[13].value
    addr3 = nom + ", France"

    # géocoder
    geolocator = Nominatim(timeout=80)
    try:
        location = geolocator.geocode(addr)
    except GeocoderTimedOut:
        print('TimeOut: Try Again')
        continue

    # tester si un résultat a été renvoyé et dégrader la précision
    if not location:
        addr = addr2
        location = geolocator.geocode(addr)
    else:
        pass

    # 2ème niveau de vérification
    if not location:
        addr = addr3
        location = geolocator.geocode(addr)
    else:
        pass

    # on print histoire de montrer ce que l'on a trouvé
    try:
        # print("adresse tentée : " + addr)
        print(location.address)
        print((location.latitude, location.longitude))
    except UnicodeEncodeError:
        # print(addr.encode("UTF-8"))
        print(location.address.encode("utf8"))
        print((location.latitude, location.longitude))

    # adding the coordinates obtained into the file
    row[33].value = location.longitude
    row[34].value = location.latitude


# ajout des coordonnées calculées par Nominatim
wb.save('ParisBeerWeek_participants.xlsx')

# ############################## EVENEMENTS

# Structure attendue ##################################
# col_idx   col_name        description
# 0         ID
# 1         EVT_NOM
# 2         TYPE
# 3         DESCR_FR
# 4         DESCR_EN
# 5         ADR_ID
# 6         ADR_NOM
# 7         ADR_PARTIC
# 8         ADR_NUM
# 9         ADR_COMPL
# 10        ADR_TYP
# 11        ADR_LIB
# 12        ADR_CP
# 13        ADR_CITY
# 14        ADR_COUNTRY
# 15        ADR_CONCAT
# 16        ADR_IMG
# 17        DDAY_START
# 18        TIME_START
# 19        DDAY_END
# 20        TIME_END
# 21        DTIME_START
# 22        DTIME_END
# 23        DUREE
# 24        OSM_URL
# 25        GMAPS_URL
# 26        DDAY_URL_FR
# 27        DDAY_URL_EN
# 28        LI_ID_PART
# 29        ED_YEAR
# 30        X_LONGITUDE
# 31        Y_LATITUDE
# 32        X_NOMINATIM
# 33        Y_NOMINATIM

# ouverture du fichier des participants en lecture
wb = load_workbook(filename='ParisBeerWeek_evenements.xlsx',
                   # read_only=True,
                   guess_types=True,
                   data_only=True,
                   # use_iterators=True
                   )

# noms des onglets
# print(wb.get_sheet_names())

ws = wb.worksheets[0]  # ws = première feuille

row_count = ws.get_highest_row() - 1
column_count = ws.get_highest_column() + 1

print(row_count)
print(column_count)

dico_lieux = {}

for row in ws.iter_rows(row_offset=1):
    # extraire le nom
    nom = row[1].value

    # vérification qu'il s'agit bien d'une ligne remplie
    if not nom:
        print('\nFin du tableau')
        break
    else:
        pass

    # on stocke les coordonnées par lieu unique pour optimiser et éviter qu'un même lieu ait 2 coordonnées différentes
    if row[5].value in dico_lieux:
        print("\tAlready localized: " + str(row[0].value))
        # on récupère les coordonnées préalablement enregistrées
        row[32].value = dico_lieux.get(row[5].value)[0]
        row[33].value = dico_lieux.get(row[5].value)[1]
        continue
    else:
        pass

    if row[13].value:
        ville = row[13].value
        addr = nom + ", " + ville + ", France"
        addr2 = row[15].value
        addr3 = nom + ", France"
    else:
        print(row[0].value)
        addr = nom + ", Île-de-France, France"
        addr2 = nom + ", France"
        addr3 = nom + ", France"

    # géocoder
    geolocator = Nominatim(timeout=80)
    try:
        location = geolocator.geocode(addr)
    except GeocoderTimedOut:
        print('TimeOut: Try Again')
        continue

    # tester si un résultat a été renvoyé et dégrader la précision
    if not location:
        addr = addr2
        location = geolocator.geocode(addr)
    else:
        pass

    # 2ème niveau de vérification
    if not location:
        addr = addr3
        location = geolocator.geocode(addr)
    else:
        pass

    # 3ème niveau de vérification
    if not location:
        print('\nAdresse NR : ' + str(row[0].value))
        continue
    else:
        pass

    # on print histoire de montrer ce que l'on a trouvé
    try:
        print(location.address)
        print((location.latitude, location.longitude))
    except UnicodeEncodeError:
        print(location.address.encode("utf8"))
        print((location.latitude, location.longitude))

    # adding the coordinates obtained into the file
    row[32].value = location.longitude
    row[33].value = location.latitude

    # on stocke les coordonnées par lieu unique pour optimiser et éviter qu'un même lieu ait 2 coordonnées différentes
    if row[5].value not in dico_lieux:
        dico_lieux[row[5].value] = (location.longitude, location.latitude)
    else:
        pass

# ajout des coordonnées calculées par Nominatim
wb.save('ParisBeerWeek_evenements.xlsx')

# #############################################################################
# ##### Stand alone program ########
# ##################################

# if __name__ == '__main__':
#     """ standalone execution """
